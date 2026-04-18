import os, sys

root = os.getcwd()
sys.path.append(os.path.join(root, 'backend'))

from database import get_all_teachers
from scheduler import generate
from export import export_excel

teachers = get_all_teachers()
timetable = generate(teachers, [])
export_excel(timetable['teachers'], timetable['classes'])

import openpyxl
wb = openpyxl.load_workbook('timetable.xlsx')
print('Sheets in generated timetable.xlsx:')
for s in wb.sheetnames:
    print(' -', s)
print('# sheets =', len(wb.sheetnames))
