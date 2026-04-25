import os
import pandas as pd
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from database import get_all_teachers

DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri"]
PERIODS = 8
PERIOD_HEADERS = [
    ("P1", "8:30 - 9:20"),
    ("P2", "9:20 - 10:10"),
    ("P3", "10:10 - 11:00"),
    ("B1", "11:00 - 11:10"),
    ("P4", "11:10 - 12:00"),
    ("P5", "12:00 - 12:50"),
    ("Lunch", "12:50 - 1:25"),
    ("P6", "1:25 - 2:15"),
    ("P7", "2:15 - 3:05"),
    ("B3", "3:05 - 3:15"),
    ("P8", "3:15 - 4:05"),
]

SHORT_ACTIVITY_LABELS = {
    'physical training': 'PT',
    'pt': 'PT',
    'mini project': 'MP',
    'professional elective': 'PE',
    'open elective': 'OE',
    'skillrack/placement': 'SR/PM',
    'skillrack placement': 'SR/PM',
    'library': 'Lib',
    'project phase / internship': 'PI',
    'project phase internship': 'PI',
}

def get_short_activity_name(name):
    if not name or not isinstance(name, str):
        return name
    return SHORT_ACTIVITY_LABELS.get(name.strip().lower(), name)


def format_teacher_slot(slot):
    if not slot or slot.get("type") is None:
        return "Free"
    if slot["type"] == "Break":
        return "Break"
    elif slot["type"] == "Free":
        return "Free"
    elif slot["type"] == "R&D":
        return "R&D"

    subject = get_short_activity_name(slot.get("subject", ""))
    class_name = slot.get("class", "")
    entry_type = slot.get("type", "Class")
    elective_no = slot.get("elective_no")

    if entry_type == "Lab":
        if class_name:
            return f"{subject}\n({class_name})\n[LAB]"
        return f"{subject}\n[LAB]"
    elif entry_type == "Activity":
        label = subject
        if class_name:
            label += f"\n({class_name})"
        if elective_no:
            label += f" -{elective_no}"
        return label
    else:
        return f"{subject}\n({class_name})" if class_name else subject


def format_class_slot(slot):
    if not slot or slot.get("type") is None:
        return "Free"
    if slot["type"] == "Break":
        return "Break"
    elif slot["type"] == "Free":
        return "Free"
    elif slot["type"] == "R&D":
        return "R&D"

    subject = get_short_activity_name(slot.get("subject", ""))
    entry_type = slot.get("type", "Class")

    if entry_type == "Lab":
        return f"{subject}\n[LAB]"
    elif entry_type == "Activity":
        if slot.get("elective_no"):
            return f"{subject} -{slot.get('elective_no')}"
        return f"{subject}\n[Activity]"
    return subject


def build_timetable_row(schedule, day, formatter):
    row = [day]
    period_index = 0
    day_schedule = schedule.get(day, []) if isinstance(schedule, dict) else []
    for label, _ in PERIOD_HEADERS:
        if label.startswith("B"):
            row.append("Break")
        elif label == "Lunch":
            row.append("Lunch")
        else:
            slot = day_schedule[period_index] if period_index < len(day_schedule) else {"type": "Free", "subject": None, "class": None}
            row.append(formatter(slot))
            period_index += 1
    return row


def export_excel(teacher_timetable, class_timetable=None, output_path="timetable.xlsx"):
    """Export timetable to Excel format with separate sheets for each teacher and class"""
    try:
        output_path = os.path.abspath(output_path)
        if os.path.exists(output_path):
            os.remove(output_path)
            print("Removed existing Excel output file.")
        print("\n=== EXCEL EXPORT START ===")
        print(f"Teachers to export: {list(teacher_timetable.keys())}")
        print(f"Total teachers: {len(teacher_timetable)}")
        print(f"Classes to export: {list(class_timetable.keys()) if class_timetable else 'None'}")
        print(f"Total classes: {len(class_timetable) if class_timetable else 0}")
        
        # Get all teachers for name lookup
        all_teachers_list = get_all_teachers()
        teacher_lookup = {str(t.teacher_id): t for t in all_teachers_list}
        print(f"Teachers in database: {list(teacher_lookup.keys())}")
        
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            # Export All Teacher Timetables in One Sheet
            exported_teachers = 0
            combined_data = []
            
            for teacher_id in sorted(teacher_timetable.keys()):
                teacher_info = teacher_lookup.get(str(teacher_id))
                teacher_name = teacher_info.name if teacher_info else f"Teacher_{teacher_id}"
                teacher_rnd_day = getattr(teacher_info, 'rnd_day', None) if teacher_info else None
                teacher_heading = f"Name : {teacher_name}"
                if teacher_rnd_day:
                    teacher_heading += f" , R&D day : {teacher_rnd_day}"

                try:
                    schedule = teacher_timetable[teacher_id]

                    # Add teacher header rows with name and R&D day
                    combined_data.append([f"Name : {teacher_name}"] + [""] * len(PERIOD_HEADERS))
                    if teacher_rnd_day:
                        combined_data.append([f"R&D day : {teacher_rnd_day}"] + [""] * len(PERIOD_HEADERS))

                    # Build header rows
                    combined_data.append(["Day"] + [label for label, _ in PERIOD_HEADERS])
                    combined_data.append(["Time"] + [time for _, time in PERIOD_HEADERS])

                    for day in DAYS:
                        combined_data.append(build_timetable_row(schedule, day, format_teacher_slot))

                    # Add blank row separator between teachers
                    combined_data.append([""] * (len(PERIOD_HEADERS) + 1))

                    print(f"Added teacher {teacher_id}: {teacher_name}")
                    exported_teachers += 1

                except Exception as e:
                    print(f"Error exporting teacher {teacher_id} ({teacher_name}): {str(e)}")
                    raise

            # Write all teachers to a single sheet
            if combined_data:
                df = pd.DataFrame(combined_data)
                df.to_excel(writer, sheet_name="Teachers Timetable", index=False, header=False)
                
                # Apply styling using openpyxl
                ws = writer.sheets["Teachers Timetable"]
                thin_border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000')
                )
                
                row_num = 1
                teacher_count = 0
                
                for idx, row in enumerate(combined_data):
                    current_row = row_num
                    
                    # Teacher heading rows
                    if row[0].startswith("Name :") or row[0].startswith("R&D day :"):
                        if row[0].startswith("Name :"):
                            teacher_count += 1
                        for col_num, value in enumerate(row, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.value = value
                            cell.border = thin_border
                            cell.font = Font(bold=True, size=12, color="000000")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            ws.column_dimensions[cell.column_letter].width = 14
                        row_num += 1
                    
                    # Header row (Day, P1, P2, ...)
                    elif row[0] == "Day":
                        for col_num, value in enumerate(row, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.value = value
                            cell.border = thin_border
                            cell.font = Font(bold=True, size=11, color="FFFFFF")
                            cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            if col_num == 1:
                                ws.column_dimensions[cell.column_letter].width = 12
                            else:
                                ws.column_dimensions[cell.column_letter].width = 14
                        row_num += 1
                    
                    # Time row
                    elif row[0] == "Time":
                        for col_num, value in enumerate(row, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.value = value
                            cell.border = thin_border
                            cell.font = Font(italic=True, size=10, color="000000")
                            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            if col_num == 1:
                                ws.column_dimensions[cell.column_letter].width = 12
                            else:
                                ws.column_dimensions[cell.column_letter].width = 14
                        row_num += 1
                    
                    # Day data rows
                    elif row[0] in DAYS:
                        for col_num, value in enumerate(row, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.value = value
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                            if col_num == 1:
                                cell.font = Font(bold=True, size=10)
                                ws.column_dimensions[cell.column_letter].width = 12
                            else:
                                ws.column_dimensions[cell.column_letter].width = 14
                                if value == 'R&D':
                                    cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
                                elif value in ('Break', 'Lunch'):
                                    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                        ws.row_dimensions[current_row].height = 45
                        row_num += 1
                    
                    # Blank separator rows
                    else:
                        row_num += 1
                
                print(f"All {exported_teachers} teachers exported to 'Teachers Timetable' sheet with formatting")
            
            # Export Student/Class Timetables - All in One Sheet
            if class_timetable and len(class_timetable) > 0:
                exported_classes = 0
                combined_class_data = []
                
                for class_name in sorted(class_timetable.keys()):
                    try:
                        schedule = class_timetable[class_name]
                        
                        # Add class name as a separator row
                        combined_class_data.append([f"CLASS: {class_name.upper()}"] + [""] * len(PERIOD_HEADERS))
                        
                        # Build header rows
                        combined_class_data.append(["Day"] + [label for label, _ in PERIOD_HEADERS])
                        combined_class_data.append(["Time"] + [time for _, time in PERIOD_HEADERS])
                        
                        for day in DAYS:
                            combined_class_data.append(build_timetable_row(schedule, day, format_class_slot))
                        
                        # Add blank row separator between classes
                        combined_class_data.append([""] * (len(PERIOD_HEADERS) + 1))
                        
                        print(f"Added class {class_name}")
                        exported_classes += 1
                        
                    except Exception as e:
                        print(f"Error exporting class {class_name}: {str(e)}")
                        raise
                
                # Write all classes to a single sheet with formatting
                if combined_class_data:
                    df = pd.DataFrame(combined_class_data)
                    df.to_excel(writer, sheet_name="Students Timetable", index=False, header=False)
                    
                    # Apply styling using openpyxl
                    ws_class = writer.sheets["Students Timetable"]
                    thin_border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000')
                    )
                    
                    row_num = 1
                    
                    for idx, row in enumerate(combined_class_data):
                        current_row = row_num
                        
                        # Class name row
                        if row[0].startswith("CLASS:"):
                            for col_num, value in enumerate(row, 1):
                                cell = ws_class.cell(row=current_row, column=col_num)
                                cell.value = value
                                cell.border = thin_border
                                cell.font = Font(bold=True, size=13, color="000000")
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                ws_class.column_dimensions[cell.column_letter].width = 14
                            row_num += 1
                        
                        # Header row (Day, P1, P2, ...)
                        elif row[0] == "Day":
                            for col_num, value in enumerate(row, 1):
                                cell = ws_class.cell(row=current_row, column=col_num)
                                cell.value = value
                                cell.border = thin_border
                                cell.font = Font(bold=True, size=11, color="FFFFFF")
                                cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                if col_num == 1:
                                    ws_class.column_dimensions[cell.column_letter].width = 12
                                else:
                                    ws_class.column_dimensions[cell.column_letter].width = 14
                            row_num += 1
                        
                        # Time row
                        elif row[0] == "Time":
                            for col_num, value in enumerate(row, 1):
                                cell = ws_class.cell(row=current_row, column=col_num)
                                cell.value = value
                                cell.border = thin_border
                                cell.font = Font(italic=True, size=10, color="000000")
                                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                if col_num == 1:
                                    ws_class.column_dimensions[cell.column_letter].width = 12
                                else:
                                    ws_class.column_dimensions[cell.column_letter].width = 14
                            row_num += 1
                        
                        # Day data rows
                        elif row[0] in DAYS:
                            for col_num, value in enumerate(row, 1):
                                cell = ws_class.cell(row=current_row, column=col_num)
                                cell.value = value
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
                                if col_num == 1:
                                    cell.font = Font(bold=True, size=10)
                                    ws_class.column_dimensions[cell.column_letter].width = 12
                                else:
                                    ws_class.column_dimensions[cell.column_letter].width = 14
                                    if value == 'R&D':
                                        cell.fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')
                                    elif value in ('Break', 'Lunch'):
                                        cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                            ws_class.row_dimensions[current_row].height = 45
                            row_num += 1
                        
                        # Blank separator rows
                        else:
                            row_num += 1
                    
                    print(f"All {exported_classes} classes exported to 'Students Timetable' sheet with formatting")
        
        print(f"Total teachers exported: {exported_teachers}")
        print("Excel file saved.")
        print("=== EXCEL EXPORT COMPLETE ===\n")
        
    except Exception as e:
        print(f"FATAL ERROR in export_excel: {str(e)}")
        raise


def export_pdf(teacher_timetable, class_timetable=None, output_path="timetable.pdf"):
    """Export timetable to PDF format with table per teacher and per class"""
    try:
        output_path = os.path.abspath(output_path)
        if os.path.exists(output_path):
            os.remove(output_path)
            print("Removed existing PDF output file.")
        print("\n=== PDF EXPORT START ===")
        print(f"Teachers to export: {list(teacher_timetable.keys())}")
        print(f"Total teachers: {len(teacher_timetable)}")
        print(f"Classes to export: {list(class_timetable.keys()) if class_timetable else 'None'}")
        print(f"Total classes: {len(class_timetable) if class_timetable else 0}")
        
        # Get all teachers for name lookup
        all_teachers_list = get_all_teachers()
        teacher_lookup = {str(t.teacher_id): t for t in all_teachers_list}
        
        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            leftMargin=0.35 * inch,
            rightMargin=0.35 * inch,
            topMargin=0.35 * inch,
            bottomMargin=0.35 * inch,
        )
        story = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=30,
            alignment=1
        )
        
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=18,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        teacher_style = ParagraphStyle(
            'TeacherName',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#003366'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        rnd_style = ParagraphStyle(
            'RndTag',
            parent=styles['Normal'],
            fontSize=11,
            backColor=colors.HexColor('#e8f4ff'),
            textColor=colors.HexColor('#0e3d7b'),
            leftIndent=0,
            rightIndent=0,
            spaceAfter=10,
            spaceBefore=0,
            borderPadding=6
        )
        
        class_style = ParagraphStyle(
            'ClassName',
            parent=styles['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#1f4788'),
            spaceAfter=10,
            spaceBefore=10
        )
        
        # Title
        story.append(Paragraph("Faculty AI Timetable Management System", title_style))
        story.append(Spacer(1, 0.3 * inch))
        
        # TEACHER TIMETABLES SECTION
        story.append(Paragraph("TEACHER TIMETABLES", section_style))
        
        exported_teachers = 0
        teacher_ids = sorted(teacher_timetable.keys())
        
        for idx, teacher_id in enumerate(teacher_ids):
            teacher_info = teacher_lookup.get(teacher_id)
            if not teacher_info:
                print(f"Warning: Teacher {teacher_id} not found in database - skipping")
                continue
            
            teacher = teacher_info
            
            try:
                schedule = teacher_timetable[teacher_id]
                
                # Teacher header - name and R&D day on separate lines
                story.append(Paragraph(f"Name : {teacher.name}", teacher_style))
                if getattr(teacher, 'rnd_day', None):
                    story.append(Paragraph(f"R&D day : {teacher.rnd_day}", rnd_style))

                # Build table data with period labels and timings
                table_data = []
                table_data.append(["Day"] + [label for label, _ in PERIOD_HEADERS])
                table_data.append(["Time"] + [time for _, time in PERIOD_HEADERS])

                r_and_d_rows = []
                for day in DAYS:
                    table_data.append(build_timetable_row(schedule, day, format_teacher_slot))
                    if "R&D" in table_data[-1][1:]:
                        r_and_d_rows.append(len(table_data) - 1)

                # Create and style table with equal column widths
                col_widths = [0.8 * inch] + [(doc.width - 0.8 * inch) / len(PERIOD_HEADERS)] * len(PERIOD_HEADERS)
                table = Table(table_data, colWidths=col_widths)
                table_style = TableStyle([
                    ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
                    ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#d9d9d9')),
                    ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),
                    ('BACKGROUND', (0, 2), (0, -1), colors.HexColor('#e6f0ff')),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ROWBACKGROUNDS', (0, 2), (-1, -1), [colors.white, colors.HexColor('#f0f5ff')]),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ])
                for row_idx in r_and_d_rows:
                    table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#FFE6CC'))
                for row_idx in range(2, len(table_data)):
                    for col_idx in range(1, len(table_data[row_idx])):
                        value = table_data[row_idx][col_idx]
                        if value in ('Break', 'Lunch'):
                            table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#FFF2CC'))
                table.setStyle(table_style)
                story.append(table)

                # Add spacing when there is more content to follow
                if idx != len(teacher_ids) - 1 or (class_timetable and len(class_timetable) > 0):
                    story.append(Spacer(1, 0.4 * inch))

                # Add page break after every 2 teachers when there are more teachers remaining
                if (idx + 1) % 2 == 0 and idx != len(teacher_ids) - 1:
                    story.append(PageBreak())
                    story.append(Paragraph("TEACHER TIMETABLES", section_style))

                print(f"Exported teacher {teacher_id}: {teacher.name}")
                exported_teachers += 1

            except Exception as e:
                print(f"Error exporting teacher {teacher_id}: {str(e)}")
                raise
        
        # STUDENT/CLASS TIMETABLES SECTION
        if class_timetable and len(class_timetable) > 0:
            story.append(Paragraph("STUDENT/CLASS TIMETABLES", section_style))
            
            exported_classes = 0
            class_names = sorted(class_timetable.keys())
            for idx, class_name in enumerate(class_names):
                try:
                    schedule = class_timetable[class_name]
                    
                    # Class info
                    class_info = f"<b>Class:</b> {class_name}"
                    story.append(Paragraph(class_info, class_style))
                    
                    # Build table data
                    table_data = []
                    table_data.append(["Day"] + [label for label, _ in PERIOD_HEADERS])
                    table_data.append(["Time"] + [time for _, time in PERIOD_HEADERS])
                    r_and_d_rows = []

                    for day in DAYS:
                        table_data.append(build_timetable_row(schedule, day, format_class_slot))
                        if "R&D" in table_data[-1][1:]:
                            r_and_d_rows.append(len(table_data) - 1)

                    # Create and style table using the same width strategy as teacher tables.
                    class_col_widths = [0.8 * inch] + [(doc.width - 0.8 * inch) / len(PERIOD_HEADERS)] * len(PERIOD_HEADERS)
                    table = Table(table_data, colWidths=class_col_widths)
                    table_style = TableStyle([
                        ('BOX', (0, 0), (-1, -1), 1.5, colors.black),
                        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#d9d9d9')),
                        ('TEXTCOLOR', (0, 1), (-1, 1), colors.black),
                        ('BACKGROUND', (0, 2), (0, -1), colors.HexColor('#f0f0f0')),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('ROWBACKGROUNDS', (0, 2), (-1, -1), [colors.white, colors.HexColor('#f9f9f9')]),
                    ])
                    for row_idx in r_and_d_rows:
                        table_style.add('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#FFE6CC'))
                    for row_idx in range(2, len(table_data)):
                        for col_idx in range(1, len(table_data[row_idx])):
                            value = table_data[row_idx][col_idx]
                            if value in ('Break', 'Lunch'):
                                table_style.add('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.HexColor('#FFF2CC'))
                    table.setStyle(table_style)
                    
                    story.append(table)

                    # Add spacing when there are more class schedules following
                    if idx != len(class_names) - 1:
                        story.append(Spacer(1, 0.4 * inch))

                    # Force two classes per PDF page when more classes remain
                    if (idx + 1) % 2 == 0 and idx != len(class_names) - 1:
                        story.append(PageBreak())
                        story.append(Paragraph("STUDENT/CLASS TIMETABLES", section_style))

                    print(f"Exported class {class_name}")
                    exported_classes += 1

                except Exception as e:
                    print(f"Error exporting class {class_name}: {str(e)}")
                    raise
            
            print(f"\nTotal classes exported: {exported_classes}")
        
        doc.build(story)
        print(f"Total teachers exported: {exported_teachers}")
        print("PDF file saved.")
        print("=== PDF EXPORT COMPLETE ===\n")
        
    except Exception as e:
        print(f"FATAL ERROR in export_pdf: {str(e)}")
        raise
